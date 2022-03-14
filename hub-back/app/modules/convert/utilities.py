#!/usr/bin/env python

"""Convert Utilities package.

For license and copyright information please see the LICENSE document (the
"License") included with this software package. This file may not be used
in any manner except in compliance with the License unless required by
applicable law or agreed to in writing, software distributed under the
License is distributed on an "AS IS" BASIS, WITHOUT WARRANTIES OR
CONDITIONS OF ANY KIND, either express or implied.

See the License for the specific language governing permissions and
limitations under the License.
"""

from flask import abort
from flask import request
from flask import json
from flask import make_response
from flask import jsonify
from flask import current_app
from urllib.parse import urlparse

from crontab import CronTab

import json
import os
import os.path

import urllib3
import csv
import pytz
import datetime

import wget

import requests
import time

import csv

from datetime import datetime
from datetime import timezone

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import Workbook

from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


from flask import current_app

def get_data_from_csv_row(csv,field_name,row):

	result = csv[row][field_name]

	return result

def import_csv(source):
	"""Import a comma separated document into the Clean Water Hub API.

	:param (object) self
		the current class (i.e., Application)

	@return (object) response
		the fully qualified response object
	"""
	result = []

	with open(source) as f:
		reader = csv.DictReader(f, delimiter=',')
		count = sum(1 for _ in reader)
		f.seek(0)
		reader = csv.DictReader(f, delimiter=',')
		for line in reader:
			result.append(line)
						
	return result

def convert_data_from_csv(source,config):
	""" START CONVERSION PROCESS """

	thing_id = ""
	location_id = ""
	data_stream_id = ""

	""" RESULT VALUE DEFINITIONS """
	result_name = ""
	result_description = ""
	restult_properties = {}
	data_streams = []
	observations = []
	observed_properties = []
	sensor = {}
	unit_of_measurements =[]
	locations = []


	output = []
	datasstreams = []
	error_log = []
	# GET NAME OF COLUMN

	# REPEAT FOR EVERY ROW IN THE SHEET


	_csv = import_csv(source)

	
	thing_name_value = config['settings']['thing_name_column']
	thing_description_value = config['settings']['thing_description_column']
	thing_lng_value = config['settings']['thing_lng_column']
	thing_lat_value = config['settings']['thing_lat_column']

	
	# print("thing name", thing_name_value)
	for i in range(0, len(_csv)):
		thing_name =  "%s" % (get_data_from_csv_row(_csv,thing_name_value,i))
		thing_description =  "%s" % (get_data_from_csv_row(_csv,thing_description_value,i))
		lng =  "%s" % (get_data_from_csv_row(_csv,thing_lng_value,i))
		lat =  "%s" % (get_data_from_csv_row(_csv,thing_lat_value,i))
		
		thing_id = thing_name.lower()
		thing_id = thing_id.replace(" ", "_")

		parameters = []


		for param in config['parameters']:
			observation_type = param['observation_type'].strip()
			property_definition = param['property_definition'].strip()
			property_description = param['property_description'].strip()
			property_name = param['property_name'].strip()
			sensor_description = param['sensor_description'].strip()
			sensor_encoding_type = param['sensor_encoding_type'].strip()
			sensor_metadata = param['sensor_metadata'].strip()
			sensor_name = param['sensor_name'].strip()
			unit_definition = param['unit_definition'].strip()
			unit_name = param['unit_name'].strip()
			unit_symbol = param['unit_symbol'].strip()

			datastrem_id = "%s_%s" % (thing_id,property_name.lower())
			datastrem_id = datastrem_id.replace(" ", "_")

			parameters.append({
				"name":thing_name,
				"description":thing_description,
				"@iot.id":datastrem_id,
				"observation_type":observation_type,
				"property_definition":property_definition,
				"property_description":property_description,
				"property_name":property_name,
				"sensor_description":sensor_description,
				"sensor_encoding_type":sensor_encoding_type,
				"sensor_metadata":sensor_metadata,
				"sensor_name":sensor_name,
				"unit_definition":unit_definition,
				"unit_name":unit_name,
				"unit_symbol":unit_symbol
			})

		for datastream in config['datastreams']:
			data_stream_property_name = datastream['name'].strip()

			data_stream_iotid = "%s_%s" % (thing_id,data_stream_property_name.lower())
			data_stream_iotid = data_stream_iotid.replace(" ", "_")
			data_stream_phenomenonTime = datastream['phenomenonTime'].strip()
			data_stream_result = datastream['result'].strip()

			data_stream_phenomenonTime = "%s" % (get_data_from_csv_row(_csv,data_stream_phenomenonTime,i)).strip()
			data_stream_result = "%s" % (get_data_from_csv_row(_csv,data_stream_result,i))

			if data_stream_result is not None and len(data_stream_result.strip()) > 0:

				if isinstance(data_stream_phenomenonTime, datetime):
			
					datasstreams.append({
						"@iot.id":data_stream_iotid,
						"phenomenonTime":data_stream_phenomenonTime.isoformat(),
						"result":data_stream_result
					})
				else: 
					if data_stream_phenomenonTime.find('/') > 0 and len(data_stream_phenomenonTime) >= 8:
						if data_stream_phenomenonTime.index('/') == 2 or data_stream_phenomenonTime.index('/') == 1:
							date_time_obj = datetime.strptime(data_stream_phenomenonTime, '%d/%m/%Y')
							datasstreams.append({
								"@iot.id":data_stream_iotid,
								"phenomenonTime":date_time_obj.isoformat(),
								"result":data_stream_result
							})
							date_time_obj = datetime.strptime(data_stream_phenomenonTime, '%d/%m/%Y')
						elif data_stream_phenomenonTime.index('/') == 4 or data_stream_phenomenonTime.index('/') == 3:
							date_time_obj = datetime.strptime(data_stream_phenomenonTime, '%Y/%m/%d')
							datasstreams.append({
								"@iot.id":data_stream_iotid,
								"phenomenonTime":date_time_obj.isoformat(),
								"result":data_stream_result
							})
						else:
							error_log.append({
								"name":thing_name,
								"error": "PhenomenonTime is not in a date format",
								"property":data_stream_property_name,
								"@iot.id":data_stream_iotid,
								"phenomenonTime":str(data_stream_phenomenonTime),
								"result":data_stream_result
							})

					elif data_stream_phenomenonTime.find('-') > 0 and len(data_stream_phenomenonTime) >= 8:
						if data_stream_phenomenonTime.index('-') == 2:
							date_time_obj = datetime.strptime(data_stream_phenomenonTime, '%d-%m-%Y')
							datasstreams.append({
								"@iot.id":data_stream_iotid,
								"phenomenonTime":date_time_obj.isoformat(),
								"result":data_stream_result
							})
						elif data_stream_phenomenonTime.index('-') == 4:

							if len(data_stream_phenomenonTime) == 24:
								datasstreams.append({
									"@iot.id":data_stream_iotid,
									"phenomenonTime":data_stream_phenomenonTime[:19],
									"result":data_stream_result
								})
							else:
								date_time_obj = datetime.strptime(data_stream_phenomenonTime[:10], '%Y-%m-%d')
								datasstreams.append({
									"@iot.id":data_stream_iotid,
									"phenomenonTime":date_time_obj.isoformat(),
									"result":data_stream_result
								})
	
						else:
							error_log.append({
								"name":thing_name,
								"error": "PhenomenonTime is not in a date format",
								"property":data_stream_property_name,
								"@iot.id":data_stream_iotid,
								"phenomenonTime":str(data_stream_phenomenonTime),
								"result":data_stream_result
							})
					else:
						error_log.append({
							
							"name":thing_name,
							"error": "PhenomenonTime is not in a date format",
							"property":data_stream_property_name,
							"@iot.id":data_stream_iotid,
							"phenomenonTime":str(data_stream_phenomenonTime),
							"result":data_stream_result
						})


		output.append({
			"@iot.id":thing_id,
			"name": thing_name,
			"description": thing_description if thing_description else "none",
			"properties": {
			},
			"lng":lng,
			"lat":lat,
			"parameters":parameters
		})

	result = {
		"status":"okay",
		"output":output,
		"datatstreams":datasstreams,
		"error_log":error_log
	}

	create_data_file_to_import(config,result)

	return result
	
def import_excel(source):

	try:
		workbook = load_workbook(filename=source)
		sheet = workbook.active
		return workbook
		
	except InvalidFileException as error:
		error_response = str(error)
		abort(make_response(jsonify(message=error_response), 400))
		return None	

def import_json(config):

	try:
		data = json.load(open(config))
		return data
	except FileNotFoundError as error:
		error_response = str(error)
		abort(make_response(jsonify(message=error_response), 400))
		return None
    
def get_data_from_excel_cell(workbook, sheet,cell):

	_sheet_reference = sheet
	_cell_reference = cell

	sheet = workbook[_sheet_reference]
	_cell = sheet[_cell_reference]

	return _cell.value

def convert_date(date):
	result = None

	if date:
		result = date

	return result.isoformat()

def get_observations(workbook,sheet,model):
	result = []
	

	for item in model:

		start = item['start']
		end = item['end']
		for i in range(start,end):

			# TODO abstract the results 0 and 1
			# TODO abstract the params
			_date = convert_date(get_data_from_excel_cell(workbook, sheet, "%s%s" % (item['phenomenonTime'],i)))
			result.append({
						
						"phenomenonTime": _date,
						"result": [
							get_data_from_excel_cell(workbook, sheet, "%s%s" % (item['result'][0],i)), 
							get_data_from_excel_cell(workbook, sheet, "%s%s" % (item['result'][1],i))
						],
						"parameters":{
							"measured_by": get_data_from_excel_cell(workbook, sheet, "%s%s" % (item['parameters']['measured_by'],i)),
							"pump": "F%s" % get_data_from_excel_cell(workbook, sheet, "%s%s" % (item['parameters']['pump'],i)),
							"comments":"G%s" % get_data_from_excel_cell(workbook, sheet, "%s%s" % (item['parameters']['comments'],i))
						}
					})
	return result

def get_data_from_sheet_or_input(mapped_to, field, workbook):
	result = ""

	if field['value_type']== 'input':
		result = field['value']
	if field['value_type']== 'sheet':
		result = get_data_from_excel_cell(workbook, \
				field['sheet'], \
				field['value'])
	return result

def convert_data_from_excel(source, config):
	""" START CONVERSION PROCESS """

	thing_id = ""
	location_id = ""
	data_stream_id = ""

	""" RESULT VALUE DEFINITIONS """
	result_name = ""
	result_description = ""
	restult_properties = {}
	data_streams = []
	observations = []
	observed_properties = []
	sensor = {}
	unit_of_measurements =[]
	locations = []

	## EXCEL CONVERTION PROCESS ##
	""" IMPORT EXECL FILE """
	workbook = import_excel(source)
	if workbook:
		sheet = workbook.active
	else:
		abort(make_response(jsonify(message='Could not import excel file'), 400))

		#Get the name

	output = []
	datasstreams = []
	error_log = []
	# GET NAME OF COLUMN

	# REPEAT FOR EVERY ROW IN THE SHEET

	sheet = config['settings']['sheet']

	sh = workbook[sheet]

	

	thing_name_value = config['settings']['thing_name_column']
	thing_description_value = config['settings']['thing_description_column']
	thing_lng_value = config['settings']['thing_lng_column']
	thing_lat_value = config['settings']['thing_lat_column']
	

	for i in range(2, sh.max_row):
		thing_name =  sh["%s%s" % (thing_name_value,i)].value
		thing_description = sh["%s%s" % (thing_description_value,i)].value
		lng = sh["%s%s" % (thing_lng_value,i)].value
		lat = sh["%s%s" % (thing_lat_value,i)].value
		thing_id = thing_name.lower()
		thing_id = thing_id.replace(" ", "_")

		parameters = []


		for param in config['parameters']:
			observation_type = param['observation_type'].strip()
			property_definition = param['property_definition'].strip()
			property_description = param['property_description'].strip()
			property_name = param['property_name'].strip()
			sensor_description = param['sensor_description'].strip()
			sensor_encoding_type = param['sensor_encoding_type'].strip()
			sensor_metadata = param['sensor_metadata'].strip()
			sensor_name = param['sensor_name'].strip()
			unit_definition = param['unit_definition'].strip()
			unit_name = param['unit_name'].strip()
			unit_symbol = param['unit_symbol'].strip()

			datastrem_id = "%s_%s" % (thing_id,property_name.lower())
			datastrem_id = datastrem_id.replace(" ", "_")

			parameters.append({
				"name":thing_name,
				"description":thing_description,
				"@iot.id":datastrem_id,
				"observation_type":observation_type,
				"property_definition":property_definition,
				"property_description":property_description,
				"property_name":property_name,
				"sensor_description":sensor_description,
				"sensor_encoding_type":sensor_encoding_type,
				"sensor_metadata":sensor_metadata,
				"sensor_name":sensor_name,
				"unit_definition":unit_definition,
				"unit_name":unit_name,
				"unit_symbol":unit_symbol
			})

		for datastream in config['datastreams']:
			data_stream_property_name = datastream['name'].strip()

			data_stream_iotid = "%s_%s" % (thing_id,data_stream_property_name.lower())
			data_stream_iotid = data_stream_iotid.replace(" ", "_")
			data_stream_phenomenonTime = datastream['phenomenonTime'].strip()
			data_stream_result = datastream['result'].strip()

			data_stream_phenomenonTime = sh["%s%s" % (data_stream_phenomenonTime,i)].value
			data_stream_result = sh["%s%s" % (data_stream_result,i)].value

			if data_stream_result is not None:
				if isinstance(data_stream_phenomenonTime, datetime):
			
					datasstreams.append({
						"@iot.id":data_stream_iotid,
						"phenomenonTime":data_stream_phenomenonTime.isoformat(),
						"result":data_stream_result
					})
				else:
					# print("#### NOT A DATE FORMAT")
					error_log.append({
						"name":thing_name,
						"error": "PhenomenonTime is not in a date format",
						"property":data_stream_property_name,
						"@iot.id":data_stream_iotid,
						"phenomenonTime":str(data_stream_phenomenonTime),
						"result":data_stream_result
					})
		

		output.append({
			"@iot.id":thing_id,
			"name": thing_name,
			"description": thing_description if thing_description else "none",
			"properties": {
			},
			"lng":lng,
			"lat":lat,
			"parameters":parameters
		})

	result = {
		"status":"okay",
		"output":output,
		"datatstreams":datasstreams,
		"error_log":error_log
	}

	create_data_file_to_import(config,result)

	return result

def convert_data(source, config):


	basepath = current_app.config['MEDIA_BASE_PATH'] + 'files/'
	directory = os.getcwd() + '/app/static/usercontent/' + 'files/'


	_source_type = None
	result = {}

	""" IMPORT JSON CONFIG FILE """
	config = import_json(directory+"/"+config)

	source = directory+source


	if config:
		pass
	else: 
		abort(make_response(jsonify(message='Could not import config file'), 400))

	""" DETECT DATA SOURCE FILE """
	if config:
		if 'settings' in config and 'type' in config['settings']:
			_source_type = config['settings']['type']

		if os.path.splitext(source)[1] == '.csv':
			_source_type = 'csv'
		elif os.path.splitext(source)[1] == '.xlsx':
			_source_type = 'excel' 
		else:
			abort(make_response(jsonify(message="File type must be .csv or .xlsx"), 400))
			
		if _source_type == 'excel':
			result = convert_data_from_excel(source, config)
		elif _source_type == 'csv':
			result = convert_data_from_csv(source,config)

	return result

def create_config(config):

	# Save JSON File
	basepath = current_app.config['MEDIA_BASE_PATH'] + 'files/'
	directory = os.getcwd() + '/app/static/usercontent/' + 'files/'

	"""
	Prepare the file for processing
	"""

	extension = "json"
	if config and 'settings' in config and 'file' in config['settings']:
		filename = config["settings"]["file"]
	else:
		filename = "config." + extension

	filepath = os.path.join(directory, filename)
	fileurl = os.path.join(basepath, filename)

	with open(filepath, 'w') as outfile:
		json.dump(config, outfile)

	return config

def create_data_file_to_import(config,data):

	# Save JSON File

	basepath = current_app.config['MEDIA_BASE_PATH'] + 'files/'
	directory = os.getcwd() + '/app/static/usercontent/' + 'files/'

	"""
	Prepare the file for processing
	"""

	extension = "json"
	if config and 'settings' in config and 'file' in config['settings']:
		filename = config["settings"]["file"]
		filename = os.path.splitext(filename)[0]+'_data_to_import.' + extension
	else:
		filename = "data_to_import." + extension

	filepath = os.path.join(directory, filename)
	fileurl = os.path.join(basepath, filename)

	with open(filepath, 'w') as outfile:
		json.dump(data, outfile)

	return data

def get_column_headers(source, file_type):

	result = {}
	sheets = []

	if file_type == "excel":

		sheet_number = 0

		workbook = import_excel(source)
		if workbook:
			sheet = workbook.active
		else:
			abort(make_response(jsonify(message='Could not import excel file'), 400))
		
		for item in workbook.sheetnames:
			sheet = workbook[item]

			list_with_values=[]
			for cell in sheet[1]:

				list_with_values.append(
					{
						'column': cell.column_letter,
						'row':cell.row,
						'value':cell.value
					}
				)

			sheets.append({
				'sheet':item,
				'sheet_number':sheet_number,
				'headers':list_with_values
			})
			sheet_number = sheet_number + 1


	elif file_type == "csv":

		_csv = import_csv(source)
		list_with_values=[]

		count = 1
		for item in _csv[0]:
			
			list_with_values.append(
					{
						'column': item,
						'row':count,
						'value':item
					}
				)
			count = count + 1
		
		sheets.append({
				'sheet':'csv',
				'sheet_number':0,
				'headers':list_with_values
			})


	result ={
			"status":"File Loaded",
			"type": "FeatureCollection",
			"features": sheets,
			"source":source
	}
	return result

def process_data(data): 

	response = None
	data_output = None

	# print("This is the data yall", data)

	data_data_stream = None
	# print("Starting Thing", "output" in data)


	if "output" in data:
		# print("Starting Things")
		for item in data["output"]:
			# print("Starting Thing")

			path = "http://frost:8080/FROST-Server/v1.1"
			thing = "Things('%s')" % item["@iot.id"]
			url_collection = "%s/Things" % (path)
			url_thing = "%s/%s" % (path, thing)
			# return url_thing



			## CHECK FOR EXISTANCE
			try:
				response = requests.get(url=url_thing,
					headers={
						"Content-Type": "application/json; charset=utf-8",
					})
				data_output = response.json()
			except requests.exceptions.RequestException:
				print('HTTP Request failed - Frost Server is not on')
				abort(make_response(jsonify(message="HTTP Request failed - Frost Server is not on"), 400))
			## END CHECK FOR EXISTANCE

			if response is None:
				abort(make_response(jsonify(message="FROST SERVER IS NOT ONLINE"), 400))
				
			if response.status_code == 200:
				continue
			elif response.status_code == 404:

				datastreams = []

				for param in item['parameters']:
					datastreams.append(
						{
						"@iot.id":param['@iot.id'],
						"name": item['name'],
						"description":item['description'],
						"observationType":param['observation_type'],
						"unitOfMeasurement": {
							"name": param['unit_name'],
							"symbol":param['unit_symbol'],
							"definition": param['unit_definition'],
						},
						"Sensor": {
							"name": param['sensor_name'],
							"description": param['sensor_description'],
							"encodingType": param['sensor_encoding_type'],
							"metadata": param['sensor_metadata'],
						},
						"ObservedProperty": {
							"name": param['property_name'],
							"definition": param['property_definition'],
							"description": param['property_description'],
						},			
					})

				data_to_post = {
					"@iot.id":item['@iot.id'],
					"name": item['name'],
					"description": item['description'],
					"Locations": [{
						"name":  item['name'],
						"description":  item['description'],
						"encodingType": "application/vnd.geo+json",
						"location": {
							"type": "Point",
							"coordinates": [item['lng'], item['lat']]
						}
					}],
					"Datastreams": datastreams
				}
				data_to_post = json.dumps(data_to_post)


				try:
					response = requests.post(url=url_collection,
						headers={
							"Content-Type": "application/json; charset=utf-8",
						},
						data=data_to_post
            			)
					# data = response.json()

					# print('Response HTTP Response Body: {content}'.format(
					# 	content=response.content))
				except requests.exceptions.RequestException:
					print('HTTP Request failed')	


	# print("#############################")
	# print("#############################")
	# print("#############################")
	# print("######## FINISHED THAT PART ")
	# print("#############################")
	# print("#############################")

	time.sleep(1)

	# print("datatstreams" in data)
	# print(len(data['datatstreams']))

	if "datatstreams" in data:
		for item in data['datatstreams']:
			# print("Starting Data Stream")

			path = "http://frost:8080/FROST-Server/v1.1"
			observation = "Datastreams('%s')/Observations" % item["@iot.id"]
			url_data_stream = "%s/%s" % (path, observation)

			## CHECK FOR EXISTANCE
			try:
				response_check = requests.get(url=url_data_stream,
					headers={
						"Content-Type": "application/json; charset=utf-8",
					})
				data_data_stream = response_check.json()
			#         print('Response HTTP Response Body: {content}'.format(
			#             content=response.content))
			except requests.exceptions.RequestException:
				print('HTTP Request failed - Frost Server is not on')
			## END CHECK FOR EXISTANCE

			data_to_post =   {
					"phenomenonTime": item["phenomenonTime"],
					"result": item["result"]
				}

			data_to_post = json.dumps(data_to_post)

			if response_check.status_code == 200:
				# POST / PATCH CHECK

				try:
					response_post_patch_check = requests.get(
						url=url_data_stream,
						params={
							"$filter": "phenomenonTime eq %s%s" % (item["phenomenonTime"],".000Z"),
						},
					)

					# IF AN OBSERVATION OF THE SAME TYPE HAS THE SAME DATE... REPLACE THE VALUE
					if response_post_patch_check.status_code == 200 and len(response_post_patch_check.json()["value"])> 0:
						print("HEY I REPLACED YOU")

						tmp_id = response_post_patch_check.json()["value"][0]["@iot.id"]
						try:
							response_to_post = requests.put(url=url_data_stream+"('"+tmp_id+"')",
								headers={
									"Content-Type": "application/json; charset=utf-8",
								},
								data=data_to_post
								)
							# data = response.json()

							# print('Response HTTP Response Body: {content}'.format(
							# 	content=response.content))
						except requests.exceptions.RequestException:
							print('HTTP Request failed')
					else:
						try:
							response_to_post = requests.post(url=url_data_stream,
								headers={
									"Content-Type": "application/json; charset=utf-8",
								},
								data=data_to_post
								)
							# data = response.json()

							# print('Response HTTP Response Body: {content}'.format(
							# 	content=response.content))
						except requests.exceptions.RequestException:
							print('HTTP Request failed')
				except requests.exceptions.RequestException:
					print('HTTP Request failed')

			elif response_check.status_code == 404:
				continue


	result = {
		"staus":"okay",
		"output":data_output,
		"datastream":data_data_stream,
		}

	return result

def run_cron(data):
	result = {}

	# 1. upload config file - write a log to json file

	# print("source", data["source"])
	# print("config", data["config"])

	if data["source"]:
		url = data["source"]
		directory = os.getcwd() + '/app/static/usercontent/' + 'files/'
		filename = wget.download(url, out=directory)

	if len(os.path.splitext(filename))>1:
		if os.path.splitext(filename)[1] == '.csv':
			_type = 'csv'
		elif os.path.splitext(filename)[1] == '.xlsx':
			_type = 'excel' 
		else:
			abort(make_response(jsonify(message="File type must be .csv or .xlsx"), 400))
	else:
		abort(make_response(jsonify(message="File type was not found"), 400))

	source_file = urlparse(data["source"])
	# print(os.path.basename(source_file.path))  # Output: 09-09-201315-47-571378756077.jpg

	# 2. get data source via url - write a log to json file
	converted_data = convert_data(os.path.basename(filename), data["config_file"])

	# 3. run prep for data upload - write a log to json file

	try:
		# print("Lets do it", "output" in converted_data)
		result = process_data(converted_data)
		# return jsonify(**result), 200
	except ValueError as e:
			abort(make_response(jsonify(message="This is not working out"), 400))


    # abort(make_response(jsonify(message="Must be in JSON format"), 400))
    # return {
    #     "content-type":request.content_type,
    #     "result":result
    #     # "data":data
    #     }

	# 4. process data - write a log to json file

	result = {
		"staus":200,
		"message":"cron Job Ran",
		"config_file":data["config_file"],
		"source":data["source"]
		}
	return result

def schedule_cron(data):

	result = {}

	current_date = datetime.now()

	if data and "cron_job_name" in data:
		cron_job_name = data["cron_job_name"]+current_date.isoformat()
	else:
		abort(make_response(jsonify(message="Cron Job Name is required"), 400))

	if data and "interval" in data:
		interval = data["interval"]
	else:
		abort(make_response(jsonify(message="Interval is required"), 400))

	# DETERMIN HOW OFTEN

	cron = CronTab(user='root')
	command = "curl 'http://localhost:5000/v1/run_job' \
	-H 'Connection: keep-alive' \
	-H 'Content-Type: application/json' \
	--data-raw '{\"config_file\":\""+data["config_file"]+"\",\"source\":\""+data["source"]+"\"}' \
	--compressed"
	job = cron.new(command=command, comment=cron_job_name)

	if interval:
		if interval == "15mins":
			job.minute.every(15)
		elif interval == "hourly":
			job.minute.every(60)
		elif interval == "nightly":
			job.hour.every(23)
			job.minute.also.on(59)
		elif interval == "weekly":
			job.dow.on('SUN')
			job.minute.also.on(59)
			job.hour.also.on(23)
		elif interval == "monthly":
			job.day.on(1)
			job.minute.also.on(59)
			job.hour.also.on(23)
			

	cron.write()
	os.system("service cron restart")
	# job.every_reboot()

	result = {
		"staus":200,
		"message":"Cron Job Scheduled",
		"config_file":data["config_file"],
		"source":data["source"]
		}

	return result

def delete_job():
	cron = CronTab(user='root')

	_name = request.args.get('name','')

	iter = cron.find_comment(_name)
	cron.remove( iter )
	cron.write()

	result = {
		"staus":201,
		"message":"A Single cron job has been removed for %s."% _name
		}

	return result

def delete_all_jobs():
	cron = CronTab(user='root')
	cron.remove_all()
	cron.write()
	result = {
		"staus":201,
		"message":"All cron jobs have been removed."
		}

	return result

def get_cron():

	jobs=[]

	cron = CronTab(user='root')
	for job in cron:
		jobs.append({
			"name":job.comment
		})
	result = {
		"staus":200,
		"jobs":jobs
		}

	return result

def get_cron_log():
	jobs = []
	result = {
		"staus":200,
		"jobs":[{
			"message":"job ran successfully"
		}]
		}

	return result